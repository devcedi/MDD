import os
import json
import re
from openpyxl import load_workbook

# CONFIGURATION 
OUTPUT_ROOT = "output"        # dossier de sortie

def get_cell_value(row, col_idx):
    # col_idx: index de colonne 1-based (ex: 114)
    # row: tuple de cellules
    return row[col_idx - 1].value if (col_idx - 1) < len(row) else None

# FONCTION utilitaire pour encoder ArtNo (slug)
def encode_filename(artno):
    # Remplace tout ce qui n'est pas alphanumérique par _
    return re.sub(r'[<>:"/\\|?*\x00-\x1F]', '', str(artno))

def get_extension(filename):
    return os.path.splitext(str(filename))[1][1:].upper() if '.' in str(filename) else ""

def CreateMDD(strName,strExelPath): 
    # Charger la configuration depuis config.json
    with open("config/"+strName+".json", "r", encoding="utf-8") as f:
        config = json.load(f)

    COLS = config["COLS"]
    ATTR_START = config["ATTR_START"]
    ATTR_END = config["ATTR_END"]


    wb = load_workbook(strExelPath)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):  # commence à la ligne 2
        artno = row[COLS["ArtNo"]-1].value
        if artno is None:
            break  # Arrêter la boucle principale si la cellule ArtNo est vide (fin des données)
        brandno = row[COLS["BrandNo"]-1].value
        brandname = row[COLS["BrandName"]-1].value
        logobrand_value = row[COLS["LogoBrand"]-1].value
        tradeno = str(row[COLS["TradeNo"]-1].value) if row[COLS["TradeNo"]-1].value else ""
        ean = str(row[COLS["EAN"]-1].value) if row[COLS["EAN"]-1].value else ""
        description = row[COLS["Description"]-1].value
        picture = row[COLS["Picture"]-1].value 
        lien = get_cell_value(row, COLS["Lien"])
        # Préparation des dossiers et noms de fichiers
        brand_folder = os.path.join(OUTPUT_ROOT, str(brandno))
        os.makedirs(brand_folder, exist_ok=True)
        json_filename = encode_filename(artno) + ".json"
        json_path = os.path.join(brand_folder, json_filename)

        # Construction du JSON
        data = {
            "ArtNo": artno,
            "Attributs": [],
            "BrandName": brandname,
            "BrandNo": str(brandno),
            "Documents": [],
            "GTIN": [],
            "GenArts": [{
                "Description": description,
                "StandardisedArticleDescription": description,
                "GenArtNo": int(brandno) * -1 if brandno else -1
            }],
            "HasPartList": False,
            "HasTextInfo": False,
            "LogoBrand": [{
                "Name": logobrand_value,
                "Extension": get_extension(logobrand_value),
                "Path": f"/img/{brandno}/{logobrand_value}"
            }] if logobrand_value else [],
            "Pictures": [{
                "Name": picture,
                "Extension": get_extension(picture),
                "Path": f"/img/{brandno}/{picture}",
                "SortNo": 1
            }] if picture else [],
            "Replaced_by": [],
            "Replaces": [],
            "Status": "",
            "TradeNo": [tradeno] if tradeno else [],
            "URL": [{"Url": lien}] if lien else []
        }

        # Ajout EAN à TradeNo si existe
        if ean:
            data["GTIN"].append(ean)

        # Traitement des Attributs
        sort_no = 1
        for j in range(ATTR_START-1, ATTR_END, 2):
            attr_name = row[j].value
            attr_value = row[j+1].value if (j+1) < len(row) else None
            if not attr_name and not attr_value:
                break
            if attr_name or attr_value:
                data["Attributs"].append({
                    "Attribute": attr_name if attr_name else "",
                    "Value": attr_value if attr_value else "",
                    "SortNo": sort_no
                })
                sort_no += 1

        # Sauvegarde du JSON
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

        print(f"Ligne {i}: fichier généré : {json_path}")

if __name__ == "__main__":
    CreateMDD("CEDILOG","INPUT/CEDILOG.xlsx")
    CreateMDD("POMMIER","INPUT/POMMIER.xlsx")
    CreateMDD("AUTOK","INPUT/AUTOK.xlsx")
    CreateMDD("PNEUMATIS","INPUT/PNEUMATIS.xlsx")
    CreateMDD("BETA","INPUT/BETA.xlsx")
    CreateMDD("CARPOLISH","INPUT/carpolish.xlsx")
    CreateMDD("DREUMEX","INPUT/DREUMEX.xlsx")        
    CreateMDD("FAABRICAUTO","INPUT/FAABRICAUTO.xlsx")
    CreateMDD("FIT","INPUT/FIT.xlsx")
    CreateMDD("KONGSBERG","INPUT/Kongsberg.xlsx")          