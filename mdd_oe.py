import json
import re
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import json
import os
import re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import json
import os
import re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def generate_artno_oe_json(
    excel_path, sheet_name=None,
    oe_start='I', oe_end='N',
    output_folder='.'
):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    start_idx = column_index_from_string(oe_start)
    end_idx = column_index_from_string(oe_end)

    for row in ws.iter_rows(min_row=2):
        artno = row[0].value
        brandno = row[1].value
        brandname = row[2].value

        if not artno:
            continue

        # Nettoyage ArtNo pour usage en nom de fichier
        pattern = r'[ <>:"/\\|?*\x00-\x1F]'
        artnoclean = re.sub(pattern, "", str(artno))
        safe_artno = artnoclean.replace("/", "_").replace("\\", "_").replace(" ", "_")

        entry = {
            "ArtNo": artno,
            "BrandName": brandname,
            "BrandNo": brandno,
            "Genart": [-brandno if isinstance(brandno, (int, float)) else None],
            "IAM": [],
            "OEM": []
        }

        all_refs = []

        # Lecture des colonnes OE uniquement (colonnes contenant les références)
        for col_idx in range(start_idx, end_idx + 1, 2):
            ref_cell = row[col_idx].value if col_idx < len(row) else None
            if ref_cell:
                all_refs.append(str(ref_cell).strip())

        if all_refs:
            entry["OEM"].append({
                "ManNo": brandno,
                "Manufacturer": "Constructeur",
                "Refs": all_refs
            })

        os.makedirs(output_folder, exist_ok=True)
        output_path = os.path.join(output_folder, f"{safe_artno}_OE.json")

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(entry, f, indent=4, ensure_ascii=False)

        print(f"✔ Fichier généré : {output_path}")




def generate_artno_manufacturer_oe_json(
    excel_path, sheet_name=None,
    oe_start='I', oe_end='N',
    output_folder='output_json'
):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    start_idx = column_index_from_string(oe_start)
    end_idx = column_index_from_string(oe_end)

    os.makedirs(output_folder, exist_ok=True)

    for row in ws.iter_rows(min_row=2):
        artno = row[0].value
        brandno = row[1].value
        brandname = row[2].value

        if not artno:
            continue

        entry = {
            "ArtNo": artno,
            "BrandName": brandname,
            "BrandNo": brandno,
            "Genart": [-brandno if isinstance(brandno, (int, float)) else None],
            "IAM": [],
            "OEM": []
        }

        oem_refs_by_manu = {}

        idx = start_idx - 1
        while idx + 1 <= end_idx - 1:
            manufacturer_cell = row[idx].value
            ref_cell = row[idx + 1].value

            if manufacturer_cell and ref_cell:
                manu = str(manufacturer_cell).strip()
                ref = str(ref_cell).strip()
                if manu in oem_refs_by_manu:
                    oem_refs_by_manu[manu].append(ref)
                else:
                    oem_refs_by_manu[manu] = [ref]

            idx += 2

        for manu, refs in oem_refs_by_manu.items():
            entry["OEM"].append({
                "ManNo": "",
                "Manufacturer": manu,
                "Refs": refs
            })

        safe_artno = str(artno).replace("/", "_").replace("\\", "_").replace(" ", "_")
        output_path = os.path.join(output_folder, f"{safe_artno}_OE.json")

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(entry, f, indent=4, ensure_ascii=False)

        print(f"✔ Fichier généré : {output_path}")



if __name__ == "__main__":
    generate_artno_oe_json(
        excel_path="INPUT/DISQUES_TAMBOURS_CEDILOG.xlsx",
        sheet_name="CROSS",
        oe_start="I",
        oe_end="I",
        output_folder="output/10000")
    
    generate_artno_manufacturer_oe_json(
        excel_path="INPUT/ETRIERS_CEDILOG.xlsx",
        sheet_name="CROSS",
        oe_start="I",
        oe_end="N",
        output_folder="output/10000")
    
    generate_artno_manufacturer_oe_json(
        excel_path="INPUT/Kongsberg.xlsx",
        sheet_name="CROSS",
        oe_start="I",
        oe_end="AF",
        output_folder="output/10021")
    
    generate_artno_oe_json(
        excel_path="INPUT/PNEUMATIS.xlsx",
        sheet_name="CROSS",
        oe_start="I",
        oe_end="BB",
        output_folder="output/10026")
    