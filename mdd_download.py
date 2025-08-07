import os
import pandas as pd
import requests
from urllib.parse import urlparse
from openpyxl import load_workbook
import urllib3
import mimetypes
from bs4 import BeautifulSoup
import os
import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib.parse import urljoin

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def download_pdf_from_page(
    url: str,
    dest_folder: str = "downloads",
    file_name: str = None
):
    try:
        # Récupérer la page HTML
        r = requests.get(url, timeout=20, verify=False)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        # Chercher un PDF dans les balises <a>, <iframe> ou <embed>
        pdf_url = None
        for tag in soup.find_all(['a', 'iframe', 'embed']):
            href = tag.get('href') or tag.get('src')
            if href and '.pdf' in href.lower():
                pdf_url = href
                break

        if not pdf_url:
            print(f"❌ Aucun lien PDF trouvé dans la page : {url}")
            return

        # Compléter l’URL relative si nécessaire
        if pdf_url.startswith("/"):
            pdf_url = urljoin(url, pdf_url)

        # Déterminer le nom du fichier
        if not file_name:
            file_name = os.path.basename(pdf_url)
            if not file_name.lower().endswith(".pdf"):
                file_name += ".pdf"

        os.makedirs(dest_folder, exist_ok=True)
        dest_path = os.path.join(dest_folder, file_name)

        print(f"🔽 Téléchargement du PDF : {pdf_url}")
        r_pdf = requests.get(pdf_url, timeout=20, verify=False)
        r_pdf.raise_for_status()

        # Obtenir le Content-Type
        content_type = r_pdf.headers.get("Content-Type", "unknown")

        with open(dest_path, "wb") as f:
            f.write(r_pdf.content)

        print(f"✅ Enregistré sous : {dest_path} | Type : {content_type}")

    except Exception as e:
        print(f"❌ Erreur pour {url} : {e}")


def download_links_from_excel_hyperlinks(
    excel_file: str,
    sheet_name: str,
    column_letter: str,  # exemple : "C"
    destination_folder: str = "downloads"
):
    os.makedirs(destination_folder, exist_ok=True)
    wb = load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name]

    for row in range(2, ws.max_row + 1):  # suppose que la première ligne est un en-tête
        # Lire la cellule contenant l'URL (ex: colonne C)
        cell = ws[f"{column_letter}{row}"]
        hyperlink = cell.hyperlink.target if cell.hyperlink else None

        # Lire la référence (colonne A)
        ref_cell = ws[f"A{row}"]
        ref_value = str(ref_cell.value).strip() if ref_cell.value else f"file_{row}"
        file_name = f"{ref_value}.pdf"

        if hyperlink and hyperlink.lower().startswith("http"):
            url = hyperlink.replace("\\", "/")  # corriger les anti-slashs
            download_pdf_from_page(
                url=url,
                dest_folder=destination_folder,
                file_basename=ref_value  # sans extension
            )
        else:
            print(f"⏭️ Ligne {row} ignorée : pas de lien valide")


def download_files_from_excel(
    excel_file: str, 
    column_index: int,   # <-- doit être un ENTIER, pas une chaîne
    destination_folder: str = "downloads"
):
    # Création du dossier si besoin
    os.makedirs(destination_folder, exist_ok=True)
    
    # Lecture du fichier Excel
    df = pd.read_excel(excel_file)
    
    # Parcours des cellules de la colonne demandée
    for idx, cell in enumerate(df.iloc[:, column_index].dropna()):
        # On ne traite que les chaînes qui ressemblent à des URLs http(s)
        if isinstance(cell, str) and cell.strip().lower().startswith("http"):
            url = cell.strip()
            try:
                # Récupère le nom du fichier à partir de l'URL
                parsed = urlparse(url)
                file_name = os.path.basename(parsed.path)
                if not file_name:
                    file_name = f"file_{idx}"
                dest_path = os.path.join(destination_folder, file_name)
                
                print(f"Downloading: {url}")
                r = requests.get(url, timeout=20, verify=False)
                r.raise_for_status()
                with open(dest_path, 'wb') as f:
                    f.write(r.content)
                print(f"  -> Saved as: {dest_path}")
            except Exception as e:
                print(f"Error downloading {url}: {e}")
        else:
            print(f"Skipped (not a URL): {cell}")

def download_pdf_from_page(
    url: str,
    dest_folder: str = "downloads",
    file_basename: str = None
):
    try:
        # Récupérer la page HTML
        r = requests.get(url, timeout=20, verify=False)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        # Chercher un document dans les balises <a>, <iframe> ou <embed>
        doc_url = None
        for tag in soup.find_all(['a', 'iframe', 'embed']):
            href = tag.get('href') or tag.get('src')
            if href and any(ext in href.lower() for ext in [".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx"]):
                doc_url = href
                break

        if not doc_url:
            print(f"❌ Aucun document trouvé dans la page : {url}")
            return

        # Compléter l’URL relative
        if doc_url.startswith("/"):
            doc_url = urljoin(url, doc_url)

        # Télécharger le document
        print(f"🔽 Téléchargement du fichier : {doc_url}")
        r_doc = requests.get(doc_url, timeout=20, verify=False)
        r_doc.raise_for_status()

        # Déterminer le type MIME
        content_type = r_doc.headers.get("Content-Type", "").split(";")[0].strip()
        extension = mimetypes.guess_extension(content_type) or ".bin"

        # Nom de fichier avec bonne extension
        if not file_basename:
            file_basename = os.path.basename(doc_url).split('.')[0] or "document"
        file_name = file_basename + extension

        os.makedirs(dest_folder, exist_ok=True)
        dest_path = os.path.join(dest_folder, file_name)

        with open(dest_path, "wb") as f:
            f.write(r_doc.content)

        print(f"✅ Enregistré sous : {dest_path} | Type : {content_type}")

    except Exception as e:
        print(f"❌ Erreur pour {url} : {e}")

if __name__ == "__main__":
    # download_files_from_excel(
    #     excel_file="INPUT/POMMIER_webmarchand TVI - info produits - template.xlsx",
    #     column_index=7,  # <-- ENTIER : colonne 8 (commence à 0)
    #     destination_folder="output/PICS/POMMIER"
    # )

    # download_links_from_excel_hyperlinks(
    #     excel_file="INPUT/ASPOCK.xlsx",
    #     sheet_name="CROSS",
    #     column_letter="H",
    #     destination_folder="output/PICS/ASPOCK"
    # )

    download_files_from_excel(
        excel_file="INPUT/FAABRICAUTO.xlsx",
        column_index=7,  # <-- ENTIER : colonne 8 (commence à 0)
        destination_folder="output/PICS/FAABRICAUTO"
    )
    # download_pdf_from_page("https://box.aspoeck.com/index.php/s/CfsXmZMyfmmiabe","output/PICS/ASPOCK","10-0100-517.pdf")